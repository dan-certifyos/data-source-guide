#!/usr/bin/env python3
"""
Regenerate the State Medicaid Exclusion Source Index tbody in certifyos-primary-source-reference-*.html
from the State Level Exclusions List.xlsx source-of-truth (column B = label, column C = hyperlink).

Usage (from repo root):
  python3 scripts/sync_medicaid_index_from_sot.py

Requires: openpyxl
  pip install openpyxl

Default xlsx path: temp-sot-downloads/State Level Exclusions List.xlsx
"""

from __future__ import annotations

import html as html_module
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "temp-sot-downloads" / "State Level Exclusions List.xlsx"
HTML_GLOB = "certifyos-primary-source-reference-*.html"
OIG_LEIE = "https://oig.hhs.gov/exclusions/exclusions_list.asp"
TERRITORIES = frozenset({"VI", "MP", "AS", "GU", "PR"})


def load_rows(xlsx: Path) -> dict[str, tuple[str, str]]:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["Sheet1"]
    rows: dict[str, tuple[str, str]] = {}
    for r in range(2, ws.max_row + 1):
        st = ws.cell(r, 1).value
        if not st or not isinstance(st, str):
            continue
        st = st.strip().upper()
        if len(st) != 2:
            continue
        name = (ws.cell(r, 2).value or "").strip().replace("\n", " ")
        cell = ws.cell(r, 3)
        url = cell.hyperlink.target if cell.hyperlink else None
        rows[st] = (name, url or "")
    return rows


def classify(us: dict[str, tuple[str, str]]):
    index_states: list[tuple[str, str, str]] = []
    footnote_states: list[str] = []
    for st in sorted(us.keys()):
        name, url = us[st]
        if not url:
            continue
        if url.rstrip("/") == OIG_LEIE or (
            url.startswith("https://oig.hhs.gov/exclusions") and "exclusions_list" in url
        ):
            footnote_states.append(st)
        else:
            index_states.append((st, name, url))
    index_states.sort(key=lambda x: x[0])
    return index_states, sorted(footnote_states)


def esc_attr(s: str) -> str:
    return html_module.escape(s, quote=True)


def esc_text(s: str) -> str:
    return html_module.escape(s, quote=False)


def build_tbody_rows(index_states: list[tuple[str, str, str]], footnote_states: list[str]) -> str:
    lines: list[str] = []
    for i in range(0, len(index_states), 2):
        left = index_states[i]
        right = index_states[i + 1] if i + 1 < len(index_states) else None
        if right:
            lines.append(
                f'<tr><td><span class="idx-state">{left[0]}</span></td>'
                f'<td><span class="idx-source"><a href="{esc_attr(left[2])}" target="_blank">{esc_text(left[1])}</a></span></td>'
                f'<td class="td-divider"></td>'
                f'<td><span class="idx-state">{right[0]}</span></td>'
                f'<td><span class="idx-source"><a href="{esc_attr(right[2])}" target="_blank">{esc_text(right[1])}</a></span></td></tr>'
            )
        else:
            lines.append(
                f'<tr><td><span class="idx-state">{left[0]}</span></td>'
                f'<td><span class="idx-source"><a href="{esc_attr(left[2])}" target="_blank">{esc_text(left[1])}</a></span></td>'
                f'<td class="td-divider"></td><td></td><td></td></tr>'
            )
    fn = ", ".join(footnote_states)
    lines.append(
        f'<tr><td colspan="2"><span style="font-size:11px;color:var(--text-muted);padding:4px 0;display:block;">'
        f'<strong>+{len(footnote_states)} jurisdictions via OIG LEIE (national):</strong> {fn} — covered by federal '
        f'<a href="https://oig.hhs.gov/exclusions/" target="_blank" style="color:var(--purple);">OIG LEIE</a>'
        f"</span></td><td class=\"td-divider\"></td><td></td><td></td></tr>"
    )
    return "\n".join(lines)


def patch_html(html_path: Path, tbody_inner: str) -> None:
    text = html_path.read_text(encoding="utf-8")
    # Replace content between State Medicaid Exclusion index <tbody> and closing before Board Licensure
    pattern = (
        r"(<!-- STATE MEDICAID EXCLUSION INDEX -->.*?"
        r"<table class=\"index-table\"[^>]*>\s*"
        r"<thead>.*?</thead>\s*<tbody>\s*)(.*?)(\s*</tbody>\s*</table>\s*"
        r"<!-- BOARD LICENSURE ACTION SOURCE INDEX -->)"
    )
    m = re.search(pattern, text, re.DOTALL)
    if not m:
        raise SystemExit(f"Could not find Medicaid index table block in {html_path}")
    new_text = m.group(1) + tbody_inner + "\n    " + m.group(3)
    text = text[: m.start()] + new_text + text[m.end() :]
    html_path.write_text(text, encoding="utf-8")


def main() -> None:
    xlsx = DEFAULT_XLSX
    if not xlsx.is_file():
        print(f"Missing {xlsx}; place the SOT export there or edit DEFAULT_XLSX.", file=sys.stderr)
        sys.exit(1)
    rows = load_rows(xlsx)
    us = {k: v for k, v in rows.items() if k not in TERRITORIES}
    index_states, footnote_states = classify(us)
    tbody = build_tbody_rows(index_states, footnote_states)
    html_files = sorted(ROOT.glob(HTML_GLOB))
    if not html_files:
        print(f"No {HTML_GLOB} in {ROOT}", file=sys.stderr)
        sys.exit(1)
    for hf in html_files:
        patch_html(hf, tbody)
        print(f"Patched {hf.name} ({len(index_states)} index rows, {len(footnote_states)} OIG footnote).")


if __name__ == "__main__":
    main()
