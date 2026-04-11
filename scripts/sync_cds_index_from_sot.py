#!/usr/bin/env python3
"""
Regenerate the CDS State Source Index tbody in certifyos-primary-source-reference.html
from 'CDS Research.xlsx' and patch the HTML file in place.

Usage (from repo root):
  python3 scripts/sync_cds_index_from_sot.py

Requires: openpyxl
  pip install openpyxl

Default xlsx path: temp-sot-downloads/CDS Research.xlsx

After running, follow the Edit → PDF → Publish steps in README.md (run publish.sh).
"""

from __future__ import annotations

import html as html_module
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "temp-sot-downloads" / "CDS Research.xlsx"
HTML_GLOB = "certifyos-primary-source-reference*.html"

# Sheets to read: (sheet_name, jur_col, board_col, url_col, dataops_col, plain_state_name)
SHEET_CONFIGS = [
    ("NurseOS", 0, 9, 7, 4, False),
    ("DentOS", 0, 10, 8, 5, False),
    ("MedOS", 0, 10, 8, 5, True),
]

# Full US state name → 2-letter code (MedOS uses plain names like "Alabama")
STATE_NAME_TO_CODE: dict[str, str] = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
    "Wisconsin": "WI", "Wyoming": "WY",
    "District of Columbia": "DC", "Washington DC": "DC", "Washington, DC": "DC",
    "Puerto Rico": "PR", "Guam": "GU",
}

BAD_URLS = {"http://link/", "link", "-", "none", ""}

# States present in the HTML but with no DataOps=Yes rows in the xlsx.
# These are included as no-URL fallback entries.
FALLBACK_ENTRIES: list[tuple[str, str]] = [
    ("AL", "Alabama Board of Medical Examiners"),
    ("DC", "DC Department of Health"),
    ("MI", "Michigan Dept. of Licensing and Regulatory Affairs"),
    ("NJ", "NJ Division of Consumer Affairs"),
    ("OK", "Oklahoma Bureau of Narcotics & Dangerous Drugs Control"),
]


def _cell_hyperlink(cell) -> str:
    return cell.hyperlink.target if cell.hyperlink else ""


def _parse_state_code(jur: str, plain_state_name: bool) -> str | None:
    """Return a 2-letter state code from a jurisdiction string."""
    jur = jur.strip()
    if not jur or jur in ("None", "#N/A") or jur.startswith("#"):
        return None
    if plain_state_name:
        return STATE_NAME_TO_CODE.get(jur)
    # Expect "Alabama (AL)" format — extract the parenthesised code.
    m = re.search(r"\(([A-Z]{2})\)", jur)
    return m.group(1) if m else None


def load_rows(xlsx: Path) -> list[tuple[str, str, str]]:
    """Return deduplicated (state_code, board_name, url) tuples from all sheets."""
    wb = load_workbook(xlsx, data_only=True)
    rows: list[tuple[str, str, str]] = []
    seen: set[tuple[str, str]] = set()

    for sheet_name, jur_col, board_col, url_col, dataops_col, plain_state_name in SHEET_CONFIGS:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found — skipping.", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        sheet_count = 0
        for row in ws.iter_rows(min_row=2):
            dataops = str(row[dataops_col].value).strip() if row[dataops_col].value else ""
            if dataops != "Yes":
                continue

            jur_raw = str(row[jur_col].value) if row[jur_col].value else ""
            state_code = _parse_state_code(jur_raw, plain_state_name)
            if not state_code:
                print(
                    f"  WARNING: [{sheet_name}] could not parse state from {jur_raw!r} — skipping.",
                    file=sys.stderr,
                )
                continue

            board = " ".join(str(row[board_col].value).split()) if row[board_col].value else ""
            if not board:
                continue

            key = (state_code, board)
            if key in seen:
                continue
            seen.add(key)

            url = _cell_hyperlink(row[url_col])
            if url.strip().lower() in BAD_URLS:
                url = ""

            rows.append((state_code, board, url))
            sheet_count += 1

        print(f"  {sheet_name}: {sheet_count} rows collected.")

    # Merge in fallback entries for states not already covered
    covered_states = {r[0] for r in rows}
    for state_code, board_name in FALLBACK_ENTRIES:
        if state_code not in covered_states:
            key = (state_code, board_name)
            if key not in seen:
                rows.append((state_code, board_name, ""))
                print(f"  Fallback added: {state_code} — {board_name}")

    rows.sort(key=lambda r: (r[0], r[1]))
    return rows


def esc_attr(s: str) -> str:
    return html_module.escape(s, quote=True)


def esc_text(s: str) -> str:
    return html_module.escape(s, quote=False)


def build_cell(state: str, board: str, url: str) -> str:
    if url:
        source_html = f'<a href="{esc_attr(url)}" target="_blank">{esc_text(board)}</a>'
    else:
        source_html = esc_text(board)
    return (
        f'<td><span class="idx-state">{esc_text(state)}</span></td>'
        f'<td><span class="idx-source">{source_html}</span></td>'
    )


def build_tbody_rows(rows: list[tuple[str, str, str]]) -> str:
    half = (len(rows) + 1) // 2
    left = rows[:half]
    right = rows[half:]

    while len(right) < len(left):
        right.append(("", "", ""))

    table_rows: list[str] = []
    for l, r in zip(left, right):
        left_cells = build_cell(*l) if l[0] else "<td></td><td></td>"
        right_cells = build_cell(*r) if r[0] else "<td></td><td></td>"
        table_rows.append(
            f'<tr>{left_cells}<td class="td-divider"></td>{right_cells}</tr>'
        )
    return "\n".join(table_rows)


def patch_html(html_path: Path, tbody_inner: str, state_count: int) -> None:
    text = html_path.read_text(encoding="utf-8")

    pattern = (
        r"(<!-- CDS SOURCE INDEX -->.*?"
        r'<table class="index-table"[^>]*>\s*'
        r"<thead>.*?</thead>\s*<tbody>\s*)(.*?)(\s*</tbody>\s*</table>\s*"
        r"<!-- EDUCATION & TRAINING -->)"
    )
    m = re.search(pattern, text, re.DOTALL)
    if not m:
        raise SystemExit(
            f"Could not find CDS Source Index table block in {html_path}"
        )
    new_text = m.group(1) + tbody_inner + "\n    " + m.group(3)
    text = text[: m.start()] + new_text + text[m.end() :]

    subtitle_pattern = r"\d+ states where CDS registration data is actively collected"
    text = re.sub(
        subtitle_pattern,
        f"{state_count} states where CDS registration data is actively collected",
        text,
        count=1,
    )

    html_path.write_text(text, encoding="utf-8")


def main() -> None:
    xlsx = DEFAULT_XLSX
    if not xlsx.is_file():
        print(
            f"Missing {xlsx}; place the SOT export there or edit DEFAULT_XLSX.",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"Reading {xlsx.name}…")
    rows = load_rows(xlsx)

    unique_states = sorted({r[0] for r in rows})
    no_url = [r for r in rows if not r[2]]
    print(
        f"\nTotal entries: {len(rows)} across {len(unique_states)} states."
    )
    if no_url:
        print(f"Entries without URLs ({len(no_url)}):")
        for state, board, _ in no_url:
            print(f"  {state}: {board}")

    tbody = build_tbody_rows(rows)

    html_files = sorted(ROOT.glob(HTML_GLOB))
    if not html_files:
        print(f"No {HTML_GLOB} in {ROOT}", file=sys.stderr)
        sys.exit(1)

    for hf in html_files:
        patch_html(hf, tbody, len(unique_states))
        print(
            f"\nPatched {hf.name} "
            f"({len(rows)} entries, {len(unique_states)} states, "
            f"{(len(rows) + 1) // 2} table rows)."
        )


if __name__ == "__main__":
    main()
