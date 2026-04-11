#!/usr/bin/env python3
"""
Regenerate the State Licensing Authority Index tbody in certifyos-primary-source-reference.html
from 'Credbase Provider Data Sources.xlsx' and patch the HTML file in place.

Usage (from repo root):
  python3 scripts/sync_state_license_index_from_sot.py

Requires: openpyxl
  pip install openpyxl

Default xlsx path: temp-sot-downloads/Credbase Provider Data Sources.xlsx

Reads the 6 per-OS detail sheets (MedOS, MentOS, DentOS, NurseOS, PT, ABA) which each
have the license verification URL as a real hyperlink in their "License Verification Link"
column. The summary "Source Log- SOT" sheet is NOT used because its col 6 XLOOKUP formulas
lose their hyperlinks when exported from Google Sheets to xlsx.

After running, follow the Edit → PDF → Publish steps in README.md (run publish.sh).
"""

from __future__ import annotations

import html as html_module
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "temp-sot-downloads" / "Credbase Provider Data Sources.xlsx"
HTML_GLOB = "certifyos-primary-source-reference*.html"

# Per-sheet config: (sheet_name, state_code_col, provider_code_col, board_name_col, url_col)
# NurseOS has a different layout from the other 5 sheets.
SHEET_CONFIGS = [
    ("MedOS ",  1, 3, 15, 13),
    ("MentOS",  1, 3, 15, 13),
    ("DentOS",  1, 3, 15, 13),
    ("NurseOS", 1, 3, 20, 18),
    ("PT ",     1, 3, 15, 13),
    ("ABA",     1, 3, 15, 13),
]

CREDENTIAL_ORDER = [
    "MD", "DO", "PA",
    "RN", "APRN", "LPN",
    "PC", "MFT", "CP", "CSW", "ADC",
    "DDS", "DMD",
    "PT",
    "ABA",
]


def _cell_hyperlink(cell) -> str:
    return cell.hyperlink.target if cell.hyperlink else ""


def _sort_credentials(creds: list[str]) -> list[str]:
    known = [c for c in CREDENTIAL_ORDER if c in creds]
    unknown = sorted(c for c in creds if c not in CREDENTIAL_ORDER)
    return known + unknown


def load_rows(xlsx: Path) -> list[tuple[str, str, str, str]]:
    """Return sorted list of (state_code, board_name, url, credential_codes_str).

    Reads all 6 per-OS detail sheets, groups by (state_code, board_name),
    collects credential codes, and resolves URLs from cell hyperlinks.
    """
    wb = load_workbook(xlsx, data_only=True)
    groups: dict[tuple[str, str], dict] = defaultdict(lambda: {"credentials": set(), "url": ""})

    for sheet_name, state_col, cred_col, board_col, url_col in SHEET_CONFIGS:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            state = str(row[state_col].value).strip() if row[state_col].value else None
            if not state or not re.fullmatch(r"[A-Z]{2}", state):
                continue

            board = " ".join(str(row[board_col].value).split()) if row[board_col].value else ""
            if not board:
                continue

            raw_creds = str(row[cred_col].value).strip() if row[cred_col].value else ""
            for code in raw_creds.splitlines():
                code = code.strip()
                if code:
                    groups[(state, board)]["credentials"].add(code)

            url = _cell_hyperlink(row[url_col])
            if url and not groups[(state, board)]["url"]:
                groups[(state, board)]["url"] = url

    rows: list[tuple[str, str, str, str]] = []
    for (state, board), entry in groups.items():
        creds_str = ", ".join(_sort_credentials(list(entry["credentials"])))
        rows.append((state, board, entry["url"], creds_str))

    rows.sort(key=lambda r: (r[0], r[1]))
    return rows


def esc_attr(s: str) -> str:
    return html_module.escape(s, quote=True)


def esc_text(s: str) -> str:
    return html_module.escape(s, quote=False)


def build_cell(state: str, board: str, url: str, creds: str) -> str:
    if url:
        source_html = f'<a href="{esc_attr(url)}" target="_blank">{esc_text(board)}</a>'
    else:
        source_html = esc_text(board)
    sub = f'<br><span class="idx-sub">{esc_text(creds)}</span>' if creds else ""
    return (
        f'<td><span class="idx-state">{esc_text(state)}</span></td>'
        f'<td><span class="idx-source">{source_html}{sub}</span></td>'
    )


def build_tbody_rows(rows: list[tuple[str, str, str, str]]) -> str:
    half = (len(rows) + 1) // 2
    left = rows[:half]
    right = rows[half:]

    while len(right) < len(left):
        right.append(("", "", "", ""))

    table_rows: list[str] = []
    for l, r in zip(left, right):
        left_cells = build_cell(*l) if l[0] else "<td></td><td></td>"
        right_cells = build_cell(*r) if r[0] else "<td></td><td></td>"
        table_rows.append(
            f'<tr>{left_cells}<td class="td-divider"></td>{right_cells}</tr>'
        )
    return "\n".join(table_rows)


def patch_html(html_path: Path, tbody_inner: str, board_count: int) -> None:
    text = html_path.read_text(encoding="utf-8")

    pattern = (
        r"(<!-- STATE LICENSING & BOARD ACTION INDEX -->.*?"
        r'<table class="index-table"[^>]*>\s*'
        r"<thead>.*?</thead>\s*<tbody>\s*)(.*?)(\s*</tbody>\s*</table>\s*"
        r"<!-- CDS SOURCE INDEX -->)"
    )
    m = re.search(pattern, text, re.DOTALL)
    if not m:
        raise SystemExit(
            f"Could not find State Licensing Authority index table block in {html_path}"
        )
    new_text = m.group(1) + tbody_inner + "\n    " + m.group(3)
    text = text[: m.start()] + new_text + text[m.end() :]

    # Update source count in subtitle (e.g. "251 sources across")
    subtitle_pattern = r"\d+ sources across"
    text = re.sub(subtitle_pattern, f"{board_count} sources across", text, count=1)

    html_path.write_text(text, encoding="utf-8")


def main() -> None:
    xlsx = DEFAULT_XLSX
    if not xlsx.is_file():
        print(
            f"Missing {xlsx}; place the SOT export there or edit DEFAULT_XLSX.",
            file=sys.stderr,
        )
        sys.exit(1)

    html_files = sorted(ROOT.glob(HTML_GLOB))
    if not html_files:
        print(f"No {HTML_GLOB} in {ROOT}", file=sys.stderr)
        sys.exit(1)

    rows = load_rows(xlsx)
    tbody = build_tbody_rows(rows)

    url_count = sum(1 for _, _, url, _ in rows if url)
    no_url_count = len(rows) - url_count

    for hf in html_files:
        patch_html(hf, tbody, len(rows))
        print(
            f"Patched {hf.name} ({len(rows)} boards, "
            f"{(len(rows) + 1) // 2} table rows, "
            f"{url_count} with URLs, {no_url_count} without)."
        )


if __name__ == "__main__":
    main()
