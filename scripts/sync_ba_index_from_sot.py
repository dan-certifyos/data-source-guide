#!/usr/bin/env python3
"""
Regenerate the Board Licensure Action Source Index tbody in certifyos-primary-source-reference.html
from 'Board Action Research - SOT.xlsx' and patch the HTML file in place.

Usage (from repo root):
  python3 scripts/sync_ba_index_from_sot.py

Requires: openpyxl
  pip install openpyxl

Default xlsx path: temp-sot-downloads/Board Action Research - SOT.xlsx

After running, follow the Edit → PDF → Publish steps in README.md (run publish.sh).
"""

from __future__ import annotations

import html as html_module
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "temp-sot-downloads" / "Board Action Research - SOT.xlsx"
HTML_GLOB = "certifyos-primary-source-reference*.html"
SHEET = "Board Action- Research"

PROCUREMENT_MAP = {
    "free": "Download",
    "view-only access; downloads not supported": "Scraper",
    "look up available": "Lookup",
    "not available": "Not available",
    "outreach": "Outreach required",
}


def map_access(proc: str) -> str:
    if not proc:
        return ""
    key = proc.strip().lower()
    for k, v in PROCUREMENT_MAP.items():
        if key.startswith(k):
            return v
    return proc.strip()


def _cell_hyperlink(cell) -> str:
    return cell.hyperlink.target if cell.hyperlink else ""


def load_rows(xlsx: Path) -> list[tuple[str, str, str, str]]:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb[SHEET]
    rows: list[tuple[str, str, str, str]] = []
    seen: set[tuple[str, str]] = set()
    for row in ws.iter_rows(min_row=2):
        state = str(row[0].value).strip() if row[0].value else None
        if not state or state in ("None", "#N/A") or state.startswith("#"):
            continue
        board = " ".join(str(row[2].value).split()) if row[2].value else ""
        proc = str(row[3].value).strip() if row[3].value else ""
        scrapable = str(row[4].value).strip() if row[4].value else ""

        if proc.lower().startswith("not available") or scrapable == "Not Scrapable":
            continue

        key = (state, board)
        if key in seen:
            continue
        seen.add(key)

        # Col M (index 12) holds the download/scraper URL; col K (index 10) holds the lookup URL.
        # Prefer col M; fall back to col K hyperlink for lookup-type boards.
        url = str(row[12].value).strip() if row[12].value else ""
        if url in ("-", "None", ""):
            url = _cell_hyperlink(row[12])
        if not url:
            url = _cell_hyperlink(row[10])

        access = map_access(proc)
        rows.append((state, board, url, access))

    rows.sort(key=lambda r: (r[0], r[1]))
    return rows


def esc_attr(s: str) -> str:
    return html_module.escape(s, quote=True)


def esc_text(s: str) -> str:
    return html_module.escape(s, quote=False)


def build_cell(state: str, board: str, url: str, access: str) -> str:
    if url:
        source_html = f'<a href="{esc_attr(url)}" target="_blank">{esc_text(board)}</a>'
    else:
        source_html = esc_text(board)
    sub = f'<span class="idx-sub">{esc_text(access)}</span>' if access else ""
    return (
        f'<td><span class="idx-state">{esc_text(state)}</span></td>'
        f'<td><span class="idx-source">{source_html}{sub}</span></td>'
    )


def build_tbody_rows(rows: list[tuple[str, str, str, str]]) -> str:
    half = (len(rows) + 1) // 2
    left = rows[:half]
    right = rows[half:]

    while len(right) < len(left):
        right.append(("", "", "", ""))

    table_rows: list[str] = []
    for l, r in zip(left, right):
        left_cells = build_cell(*l) if l[0] else "<td></td><td></td>"
        right_cells = build_cell(*r) if r[0] else "<td></td><td></td>"
        table_rows.append(
            f'<tr>{left_cells}<td class="td-divider"></td>{right_cells}</tr>'
        )
    return "\n".join(table_rows)


def patch_html(html_path: Path, tbody_inner: str, board_count: int) -> None:
    text = html_path.read_text(encoding="utf-8")

    # Replace tbody content between the board licensure action index and the provider attestation section
    pattern = (
        r"(<!-- BOARD LICENSURE ACTION SOURCE INDEX -->.*?"
        r'<table class="index-table"[^>]*>\s*'
        r"<thead>.*?</thead>\s*<tbody>\s*)(.*?)(\s*</tbody>\s*</table>\s*"
        r"<!-- PROVIDER ATTESTATION -->)"
    )
    m = re.search(pattern, text, re.DOTALL)
    if not m:
        raise SystemExit(f"Could not find Board Licensure Action index table block in {html_path}")
    new_text = m.group(1) + tbody_inner + "\n    " + m.group(3)
    text = text[: m.start()] + new_text + text[m.end() :]

    # Update the section subtitle board count
    subtitle_pattern = r'(\d+) board sources'
    text = re.sub(subtitle_pattern, f'{board_count} board sources', text, count=1)

    html_path.write_text(text, encoding="utf-8")


def main() -> None:
    xlsx = DEFAULT_XLSX
    if not xlsx.is_file():
        print(f"Missing {xlsx}; place the SOT export there or edit DEFAULT_XLSX.", file=sys.stderr)
        sys.exit(1)

    rows = load_rows(xlsx)
    tbody = build_tbody_rows(rows)

    html_files = sorted(ROOT.glob(HTML_GLOB))
    if not html_files:
        print(f"No {HTML_GLOB} in {ROOT}", file=sys.stderr)
        sys.exit(1)

    for hf in html_files:
        patch_html(hf, tbody, len(rows))
        print(f"Patched {hf.name} ({len(rows)} boards, {(len(rows) + 1) // 2} table rows).")


if __name__ == "__main__":
    main()
